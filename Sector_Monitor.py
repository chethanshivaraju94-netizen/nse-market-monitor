import yfinance as yf
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

print("--- NSE Sector Rotation Engine Initiated ---")

file_name = "NSE_Sector_Monitor.xlsx"
benchmark_ticker = "^CRSLDX" # Nifty 500

tickers = {
    # Broad & Sectoral
    "IT": "ITBEES.NS",               
    "Bank": "BANKBEES.NS",           
    "Private Bank": "HDFCPVTBAN.NS", 
    "PSU Bank": "PSUBNKBEES.NS",     
    "Auto": "AUTOBEES.NS",           
    "FMCG": "FMCGIETF.NS",           
    "Pharma": "PHARMABEES.NS",       
    "Healthcare": "HEALTHY.NS",      
    "Metal": "METALIETF.NS",         
    "Realty": "MOREALTY.NS",         
    "Fin Services": "FINIETF.NS",    
    "Infrastructure": "INFRAIETF.NS",
    "Consumption": "CONSUMBEES.NS",  
    "PSE": "ABSLPSE.NS",             
    "Energy": "MOENERGY.NS",         
    "Commodities": "COMMOIETF.NS",   
    "MNC": "MNC.NS",
    
    # Thematic & Niche
    "Defence": "MODEFENCE.NS",       
    "Oil & Gas": "OILIETF.NS",       
    "Chemicals": "CHEMICAL.NS",      
    "Manufacturing": "MAKEINDIA.NS",
    "Capital Market": "MOCAPITAL.NS",
    "Digital": "TNIDETF.NS",         
    "Internet": "INTERNET.NS",        
    "Tourism": "MOTOUR.NS",           
    "Services": "MOSERVICE.NS",       
    "EV & New Age Auto": "GROWWEV.NS" 
}

# 1. Download Data
print("Downloading 2-Year Market Data...")
all_tickers = list(tickers.values()) + [benchmark_ticker]
raw_data = yf.download(all_tickers, period="2y", group_by='ticker')

# Check for multi-index formatting from yfinance
if isinstance(raw_data.columns, pd.MultiIndex):
    close_data = pd.DataFrame()
    for t in all_tickers:
        if t in raw_data:
            close_data[t] = raw_data[t]['Close']
else:
    close_data = raw_data['Close']

# --- QUOTE PATCH: Fix YF's lagging chart API using real-time overview quotes ---
print("Patching delayed Yahoo Finance charts with live overview quotes...")
live_prices = {}
for symbol in all_tickers:
    try:
        t = yf.Ticker(symbol)
        try:
            price = t.fast_info.last_price
        except AttributeError:
            price = t.info.get('regularMarketPrice', np.nan)
            
        if pd.notna(price):
            live_prices[symbol] = float(price)
    except Exception:
        pass

if live_prices:
    last_row_nans = close_data.iloc[-1].isna().sum()
    
    if last_row_nans > 0:
        # The chart created today's row, but some indices are missing candles. Overwrite them.
        for symbol, price in live_prices.items():
            if symbol in close_data.columns:
                close_data.loc[close_data.index[-1], symbol] = price
    else:
        # The chart has no NaNs, meaning it might be fully stuck on yesterday. Check benchmark drift.
        bench_live = live_prices.get(benchmark_ticker)
        bench_chart = close_data[benchmark_ticker].iloc[-1] if benchmark_ticker in close_data.columns else None
        
        # If the live quote deviates significantly from the chart, append a new daily row.
        if bench_live and pd.notna(bench_chart) and abs(bench_live - bench_chart) > 2.0:
            print("Appending missing daily candle...")
            new_idx = close_data.index[-1] + pd.Timedelta(days=1)
            close_data.loc[new_idx] = np.nan
            for symbol, price in live_prices.items():
                if symbol in close_data.columns:
                    close_data.loc[new_idx, symbol] = price
        else:
            # The chart is fully up to date, just enforce precision from the quote API
            for symbol, price in live_prices.items():
                if symbol in close_data.columns:
                    close_data.loc[close_data.index[-1], symbol] = price

# Now apply forward fill for older historical gaps
close_data = close_data.ffill().dropna(how='all')
benchmark_close = close_data[benchmark_ticker]
# -------------------------------------------------------------------------------

# 2. Calculate Sector Matrices
print("Calculating Relative Strength and Moving Averages...")
metrics = []
historical_ranks = pd.DataFrame(index=close_data.index, columns=tickers.keys())

# Calculate daily ranks for Tab 2
for date_idx in range(65, len(close_data)):
    daily_rs_returns = {}
    for sector_name, symbol in tickers.items():
        if symbol not in close_data.columns: continue
        
        sector_series = close_data[symbol].iloc[:date_idx+1]
        bench_series = benchmark_close.iloc[:date_idx+1]
        
        # RS Line = Sector Price / Benchmark Price
        rs_line = sector_series / bench_series
        
        # 65-Day RS Momentum
        if len(rs_line) > 65:
            rs_65d = ((rs_line.iloc[-1] - rs_line.iloc[-66]) / rs_line.iloc[-66]) * 100
            daily_rs_returns[sector_name] = rs_65d
            
    # Rank them for this specific day (Highest RS % = Rank 1)
    if daily_rs_returns:
        sorted_ranks = sorted(daily_rs_returns.items(), key=lambda x: x[1], reverse=True)
        for rank, (sec, val) in enumerate(sorted_ranks, 1):
            historical_ranks.at[close_data.index[date_idx], sec] = rank

# Today's Calculations for Tab 1
for sector_name, symbol in tickers.items():
    if symbol not in close_data.columns: 
        print(f"Warning: {sector_name} ({symbol}) data unavailable.")
        continue
        
    sec_close = close_data[symbol]
    rs_line = sec_close / benchmark_close
    
    # Base Price Metrics
    chg_1d = ((sec_close.iloc[-1] - sec_close.iloc[-2]) / sec_close.iloc[-2]) * 100
    
    # Moving Averages
    ema_10 = sec_close.ewm(span=10, adjust=False).mean().iloc[-1]
    ema_20 = sec_close.ewm(span=20, adjust=False).mean().iloc[-1]
    sma_50 = sec_close.rolling(50).mean().iloc[-1]
    sma_200 = sec_close.rolling(200).mean().iloc[-1]
    
    gt_10 = "Yes" if sec_close.iloc[-1] > ema_10 else "No"
    gt_20 = "Yes" if sec_close.iloc[-1] > ema_20 else "No"
    gt_50 = "Yes" if sec_close.iloc[-1] > sma_50 else "No"
    gt_200 = "Yes" if sec_close.iloc[-1] > sma_200 else "No"
    
    # RS Metrics
    rs_5d = ((rs_line.iloc[-1] - rs_line.iloc[-6]) / rs_line.iloc[-6]) * 100 if len(rs_line) > 5 else 0
    rs_21d = ((rs_line.iloc[-1] - rs_line.iloc[-22]) / rs_line.iloc[-22]) * 100 if len(rs_line) > 21 else 0
    rs_65d = ((rs_line.iloc[-1] - rs_line.iloc[-66]) / rs_line.iloc[-66]) * 100 if len(rs_line) > 65 else 0
    
    rs_sma_50 = rs_line.rolling(50).mean().iloc[-1]
    rs_trend = "Up" if rs_line.iloc[-1] > rs_sma_50 else "Down"
    
    # Continuous % Off RS High calculation
    rs_252_max = rs_line.rolling(252).max().iloc[-1]
    pct_off_rs_high = ((rs_line.iloc[-1] - rs_252_max) / rs_252_max) * 100 if rs_252_max > 0 else 0.0
    
    # Rank Velocity (Momentum Acceleration)
    current_rank = historical_ranks[sector_name].iloc[-1]
    past_rank_5d = historical_ranks[sector_name].iloc[-6] if len(historical_ranks) >= 6 else np.nan
    past_rank_10d = historical_ranks[sector_name].iloc[-11] if len(historical_ranks) >= 11 else np.nan
    past_rank_20d = historical_ranks[sector_name].iloc[-21] if len(historical_ranks) >= 21 else np.nan
    past_rank_65d = historical_ranks[sector_name].iloc[-66] if len(historical_ranks) >= 66 else np.nan
    
    def calc_vel(past, curr):
        return int(past) - int(curr) if pd.notna(past) and pd.notna(curr) else 0

    metrics.append({
        "Sector": sector_name,
        "65D RS Rank": current_rank,
        "5D Rank Velocity": calc_vel(past_rank_5d, current_rank),
        "10D Rank Velocity": calc_vel(past_rank_10d, current_rank),
        "20D Rank Velocity": calc_vel(past_rank_20d, current_rank),
        "65D Rank Velocity": calc_vel(past_rank_65d, current_rank),
        "Close": round(sec_close.iloc[-1], 2),
        "% Chg": round(chg_1d, 2),
        "5D RS %": round(rs_5d, 2),
        "21D RS %": round(rs_21d, 2),
        "65D RS %": round(rs_65d, 2),
        "RS Trend (>50 SMA)": rs_trend,
        "% Off RS High": round(pct_off_rs_high, 2),
        "> 10 EMA": gt_10,
        "> 20 EMA": gt_20,
        "> 50 SMA": gt_50,
        "> 200 SMA": gt_200
    })

df_heatmap = pd.DataFrame(metrics).sort_values(by="65D RS Rank", ascending=True)

# 3. Build Excel File
print("Writing Engine to Excel...")
wb = Workbook()

# Tab 1: Cross-Sectional Heatmap
ws1 = wb.active
ws1.title = "Heatmap"
headers_t1 = df_heatmap.columns.tolist()
ws1.append(headers_t1)

for r in df_heatmap.itertuples(index=False):
    ws1.append(list(r))

# Tab 2: Historical Ranks Tracker
ws2 = wb.create_sheet(title="Rotation Tracker")
hist_tracker = historical_ranks.dropna(how='all').tail(65).sort_index(ascending=False)
headers_t2 = ["Date"] + list(hist_tracker.columns)
ws2.append(headers_t2)

for date, row in hist_tracker.iterrows():
    r_data = [date.strftime("%Y-%m-%d")] + [row[col] for col in hist_tracker.columns]
    ws2.append(r_data)

# 4. Apply UI Formatting
green_fill = PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid")
red_fill = PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")

# Tab 1 Formatting (Dynamic Auto-Fit to Maximum Value Length)
for col in ws1.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_let = col[0].column_letter
    ws1.column_dimensions[col_let].width = max(max_len + 4, 12) # Padding for clear visual spacing
    for cell in col: cell.alignment = center_align

# Conditional formatting for Rank Velocity (Columns C, D, E, F)
velocity_scale = ColorScaleRule(start_type='num', start_value=-10, start_color='F8696B', mid_type='num', mid_value=0, mid_color='FFFFFF', end_type='num', end_value=10, end_color='63BE7B')
ws1.conditional_formatting.add(f"C2:F{ws1.max_row}", velocity_scale)

# Apply explicit plus/minus number formatting to the velocity columns so "+5" renders cleanly
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=3, max_col=6):
    for cell in row:
        cell.number_format = '+0;-0;0'

# Conditional formatting for Time-Frame Matrices (Columns I, J, K - 5D, 21D, 65D RS)
rs_scale = ColorScaleRule(start_type='num', start_value=-10, start_color='F8696B', mid_type='num', mid_value=0, mid_color='FFFFFF', end_type='num', end_value=10, end_color='63BE7B')
ws1.conditional_formatting.add(f"I2:K{ws1.max_row}", rs_scale)

# Conditional Formatting for Binary Indicators
ws1.conditional_formatting.add(f"L2:L{ws1.max_row}", CellIsRule(operator='equal', formula=['"Up"'], fill=green_fill))
ws1.conditional_formatting.add(f"L2:L{ws1.max_row}", CellIsRule(operator='equal', formula=['"Down"'], fill=red_fill))

# Continuous Formatting for % Off RS High (Column M)
pct_off_scale = ColorScaleRule(start_type='num', start_value=-15.0, start_color='F8696B', mid_type='num', mid_value=-5.0, mid_color='FFFFFF', end_type='num', end_value=0.0, end_color='63BE7B')
ws1.conditional_formatting.add(f"M2:M{ws1.max_row}", pct_off_scale)

for col_let in ['N', 'O', 'P', 'Q']: # Moving Averages
    ws1.conditional_formatting.add(f"{col_let}2:{col_let}{ws1.max_row}", CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    ws1.conditional_formatting.add(f"{col_let}2:{col_let}{ws1.max_row}", CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))

# Tab 2 Formatting (Dynamic Auto-Fit to Maximum Value Length)
for col in ws2.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_let = col[0].column_letter
    ws2.column_dimensions[col_let].width = max(max_len + 4, 12) # Prevents dates and headers from clipping
    for cell in col: cell.alignment = center_align

# Dynamically adjusted scale for varying number of total sectors
num_sectors = len(tickers)
mid_rank = (num_sectors // 2) + 1
rank_scale = ColorScaleRule(start_type='num', start_value=1, start_color='63BE7B', mid_type='num', mid_value=mid_rank, mid_color='FFFFFF', end_type='num', end_value=num_sectors, end_color='F8696B')

last_col_letter = ws2.cell(row=1, column=ws2.max_column).column_letter
ws2.conditional_formatting.add(f"B2:{last_col_letter}{ws2.max_row}", rank_scale)

wb.save(file_name)
print(f"--- SUCCESS! Engine saved to {file_name} ---")
