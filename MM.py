import yfinance as yf
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill

print("--- NSE Market Monitor (Final VCP Master Engine) ---")
file_name = "NSE_Market_Monitor.xlsx"

# Custom Python Color Engine for Nifty 500 Drawdown
def get_drawdown_color(pct):
    if pct == "" or pd.isna(pct): return "FFFFFF"
    try:
        pct = float(pct)
    except:
        return "FFFFFF"
        
    if pct >= 0: return "63BE7B"       # New High (Max Green)
    elif pct <= -15: return "F8696B"   # Deep Correction (Max Red)
    elif pct > -5:
        # Fades from Green (0%) to White (-5%)
        ratio = abs(pct) / 5.0
        r = int(99 + (255 - 99) * ratio)
        g = int(190 + (255 - 190) * ratio)
        b = int(123 + (255 - 123) * ratio)
    else:
        # Fades from White (-5%) to Red (-15%)
        ratio = (abs(pct) - 5.0) / 10.0
        r = int(255 - (255 - 248) * ratio)
        g = int(255 - (255 - 105) * ratio)
        b = int(255 - (255 - 107) * ratio)
    return f"{r:02X}{g:02X}{b:02X}"

# 1. Define the upgraded layout with Benchmark Columns
headers = [
    "Date", "Nifty 500 Close", "Nifty 500 Chg %", 
    "Up 4% Today", "Down 4% Today", "5 Day Ratio", "10 Day Ratio", 
    "Advances", "Declines", "A/D Ratio", "52W Highs", "52W Lows", "Volume Breadth",
    "> 200 SMA (%)", "> 50 SMA (%)", "> 20 EMA (%)", "> 10 EMA (%)"
]

# 2. Fetch Nifty 500 Benchmark Data (^CRSLDX is the Yahoo ticker for Nifty 500)
print("Fetching Nifty 500 Benchmark data...")
n500_dict = {}
try:
    nifty500_raw = yf.download('^CRSLDX', period="2y")
    
    # THE FIX: Force flatten the dataframe if yfinance returns a MultiIndex
    if isinstance(nifty500_raw.columns, pd.MultiIndex):
        nifty500_raw.columns = nifty500_raw.columns.droplevel(1)
        nifty500_raw = nifty500_raw.loc[:,~nifty500_raw.columns.duplicated()]
        
    nifty500_raw['Change'] = nifty500_raw['Close'].pct_change() * 100
    nifty500_raw['52W_High'] = nifty500_raw['High'].rolling(window=252).max()
    nifty500_raw['Pct_Off_High'] = ((nifty500_raw['Close'] - nifty500_raw['52W_High']) / nifty500_raw['52W_High']) * 100
    
    for ts, r in nifty500_raw.iterrows():
        d_str = ts.strftime("%Y-%m-%d")
        
        # Safely extract floats directly
        try:
            c_val = float(r['Close'])
            chg_val = float(r['Change'])
            pct_val = float(r['Pct_Off_High'])
        except:
            c_val, chg_val, pct_val = None, None, None

        n500_dict[d_str] = {
            'Close': round(c_val, 2) if pd.notna(c_val) else "",
            'Change': round(chg_val, 2) if pd.notna(chg_val) else "",
            'Pct_Off': round(pct_val, 2) if pd.notna(pct_val) else ""
        }
except Exception as e:
    print(f"Warning: Could not fetch Nifty 500 data: {e}")

# 3. Fetch the Nifty Total Market universe list
print("Fetching Nifty Total Market universe list...")
try:
    url = "https://archives.nseindia.com/content/indices/ind_niftytotalmarket_list.csv"
    nifty_total_df = pd.read_csv(url, storage_options={'User-Agent': 'Mozilla/5.0'})
    tickers = [str(symbol) + ".NS" for symbol in nifty_total_df['Symbol']]
except Exception as e:
    print(f"Error fetching live list: {e}. Falling back to standard list.")
    tickers = ["RELIANCE.NS", "TCS.NS", "INFY.NS", "HDFCBANK.NS", "ICICIBANK.NS", "SBIN.NS"]

# 4. Download full OHLCV data for Breadth Metrics
print("Downloading breadth data from Yahoo Finance (this may take 1-2 minutes)...")
raw_data = yf.download(tickers, period="2y")

close_df = raw_data['Close'].dropna(how='all', axis=1)
valid_tickers = close_df.columns
high_df = raw_data['High'][valid_tickers]
low_df = raw_data['Low'][valid_tickers]
vol_df = raw_data['Volume'][valid_tickers]

print("Calculating daily breadth, volume, and technical matrices...")
daily_returns_df = close_df.pct_change() * 100
sma_200_df = close_df.rolling(window=200).mean()
sma_50_df = close_df.rolling(window=50).mean()
ema_20_df = close_df.ewm(span=20, adjust=False).mean()
ema_10_df = close_df.ewm(span=10, adjust=False).mean()

rolling_high_252 = high_df.rolling(window=252).max()
rolling_low_252 = low_df.rolling(window=252).min()

history_data = []
lookback_days = 65 
if len(close_df) < lookback_days: lookback_days = len(close_df) - 1

for i in range(-lookback_days, 0):
    date_str = close_df.index[i].strftime("%Y-%m-%d")
    day_close = close_df.iloc[i]
    day_returns = daily_returns_df.iloc[i]
    day_high = high_df.iloc[i]
    day_low = low_df.iloc[i]
    day_vol = vol_df.iloc[i]
    
    up_4 = int((day_returns >= 4.0).sum())
    down_4 = int((day_returns <= -4.0).sum())
    advances = int((day_returns > 0).sum())
    declines = int((day_returns < 0).sum())
    ad_ratio = round((advances / declines), 2) if declines > 0 else float(advances)
    
    new_highs = int((day_high >= rolling_high_252.iloc[i]).sum())
    new_lows = int((day_low <= rolling_low_252.iloc[i]).sum())
    
    up_vol = day_vol[day_returns > 0].sum()
    down_vol = day_vol[day_returns < 0].sum()
    if down_vol > 0:
        vol_breadth = round((up_vol / down_vol), 2)
    elif up_vol > 0:
        vol_breadth = 99.99 
    else:
        vol_breadth = 0.0
    
    sma_200 = sma_200_df.iloc[i]
    valid_200 = sma_200.notna().sum() 
    pct_200 = round(((day_close > sma_200).sum() / valid_200) * 100, 2) if valid_200 > 0 else 0.0

    sma_50 = sma_50_df.iloc[i]
    valid_50 = sma_50.notna().sum()
    pct_50 = round(((day_close > sma_50).sum() / valid_50) * 100, 2) if valid_50 > 0 else 0.0

    ema_20 = ema_20_df.iloc[i]
    valid_20 = ema_20.notna().sum()
    pct_20 = round(((day_close > ema_20).sum() / valid_20) * 100, 2) if valid_20 > 0 else 0.0

    ema_10 = ema_10_df.iloc[i]
    valid_10 = ema_10.notna().sum()
    pct_10 = round(((day_close > ema_10).sum() / valid_10) * 100, 2) if valid_10 > 0 else 0.0
    
    history_data.append({
        "Date": date_str, "Nifty 500 Close": "", "Nifty 500 Chg %": "", "Hidden_Pct_Off": "",
        "Up 4% Today": up_4, "Down 4% Today": down_4, 
        "Advances": advances, "Declines": declines, "A/D Ratio": ad_ratio, 
        "52W Highs": new_highs, "52W Lows": new_lows, "Volume Breadth": vol_breadth,
        "> 200 SMA (%)": pct_200, "> 50 SMA (%)": pct_50, "> 20 EMA (%)": pct_20, "> 10 EMA (%)": pct_10
    })

df_recent = pd.DataFrame(history_data)
df_recent['5 Day Ratio'] = (df_recent['Up 4% Today'].rolling(5).sum() / df_recent['Down 4% Today'].rolling(5).sum().replace(0, 1)).round(2)
df_recent['10 Day Ratio'] = (df_recent['Up 4% Today'].rolling(10).sum() / df_recent['Down 4% Today'].rolling(10).sum().replace(0, 1)).round(2)
df_recent = df_recent.sort_values(by="Date", ascending=False)

# 5. Smart Merge & Benchmark Backfill
print("Merging with archive and backfilling Nifty 500 Data...")
if os.path.exists(file_name):
    try:
        df_archive = pd.read_excel(file_name)
        df_combined = pd.concat([df_recent, df_archive], ignore_index=True)
        df_combined = df_combined.drop_duplicates(subset=['Date'], keep='first')
        
        for col in headers:
            if col not in df_combined.columns:
                df_combined[col] = "" 
        if 'Hidden_Pct_Off' not in df_combined.columns:
            df_combined['Hidden_Pct_Off'] = ""
            
    except Exception as e:
        print(f"Archive read error: {e}")
        df_combined = df_recent.copy()
else:
    df_combined = df_recent.copy()

# Backfill the exact index data into every row seamlessly
for idx, row in df_combined.iterrows():
    d_str = str(row['Date'])[:10]
    if d_str in n500_dict:
        df_combined.at[idx, 'Nifty 500 Close'] = n500_dict[d_str]['Close']
        df_combined.at[idx, 'Nifty 500 Chg %'] = n500_dict[d_str]['Change']
        df_combined.at[idx, 'Hidden_Pct_Off'] = n500_dict[d_str]['Pct_Off']

df_final = df_combined.sort_values(by="Date", ascending=False)
df_final = df_final.fillna("") 

# 6. Rebuild Excel File
if os.path.exists(file_name):
    os.remove(file_name)

wb = Workbook()
ws = wb.active
ws.title = "Market Monitor"

ws.append(headers)

for index, row in df_final.iterrows():
    date_val = str(row['Date'])[:10] if str(row['Date']) != "" else ""
    pct_off = row.get('Hidden_Pct_Off', "")
    
    ws.append([
        date_val, row['Nifty 500 Close'], row['Nifty 500 Chg %'],
        row['Up 4% Today'], row['Down 4% Today'], row['5 Day Ratio'], row['10 Day Ratio'],
        row['Advances'], row['Declines'], row['A/D Ratio'], 
        row['52W Highs'], row['52W Lows'], row['Volume Breadth'],
        row['> 200 SMA (%)'], row['> 50 SMA (%)'], row['> 20 EMA (%)'], row['> 10 EMA (%)']
    ])
    
    # 7A. Apply Custom Python Paint directly to the Nifty 500 Price (Column B)
    current_row = ws.max_row
    color_hex = get_drawdown_color(pct_off)
    ws[f"B{current_row}"].fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

# 7B. Apply Fixed Quantitative Conditional Formatting to everything else
max_row = ws.max_row
ws.conditional_formatting._cf_rules = {}

n500_chg_scale = ColorScaleRule(start_type='num', start_value=-2.0, start_color='F8696B', mid_type='num', mid_value=0, mid_color='FFFFFF', end_type='num', end_value=2.0, end_color='63BE7B')
thrust_green = ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF', end_type='num', end_value=200, end_color='63BE7B')
thrust_red = ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF', end_type='num', end_value=200, end_color='F8696B')
breadth_green = ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF', end_type='num', end_value=750, end_color='63BE7B')
breadth_red = ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF', end_type='num', end_value=750, end_color='F8696B')
ratio_scale = ColorScaleRule(start_type='num', start_value=0.5, start_color='F8696B', mid_type='num', mid_value=1.0, mid_color='FFFFFF', end_type='num', end_value=2.0, end_color='63BE7B')
ma_scale = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B', mid_type='num', mid_value=50, mid_color='FFFFFF', end_type='num', end_value=100, end_color='63BE7B')

ws.conditional_formatting.add(f"C2:C{max_row}", n500_chg_scale)# Nifty 500 Chg % (-2% to 2%)
ws.conditional_formatting.add(f"D2:D{max_row}", thrust_green)  # Up 4% 
ws.conditional_formatting.add(f"E2:E{max_row}", thrust_red)    # Down 4% 
ws.conditional_formatting.add(f"F2:G{max_row}", ratio_scale)   # 5 & 10 Day Ratios 
ws.conditional_formatting.add(f"H2:H{max_row}", breadth_green) # Advances 
ws.conditional_formatting.add(f"I2:I{max_row}", breadth_red)   # Declines 
ws.conditional_formatting.add(f"J2:J{max_row}", ratio_scale)   # A/D Ratio 
ws.conditional_formatting.add(f"K2:K{max_row}", thrust_green)  # 52W Highs 
ws.conditional_formatting.add(f"L2:L{max_row}", thrust_red)    # 52W Lows
ws.conditional_formatting.add(f"M2:M{max_row}", ratio_scale)   # Volume Breadth 
ws.conditional_formatting.add(f"N2:Q{max_row}", ma_scale)      # Moving Averages 

# 8. Auto-fit Columns
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter 
    for cell in col:
        try: 
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) 
    ws.column_dimensions[column].width = adjusted_width

wb.save(file_name)
print(f"\n--- SUCCESS! Final VCP Engine saved to {file_name}. ---")
